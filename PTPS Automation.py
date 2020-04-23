# -*- coding: utf-8 -*-
"""
Created on Tue Mar 10 09:34:46 2020

@author: SahilPatil
"""

from openpyxl import load_workbook
import pandas as pd

#Load PTPS Workbook
MRxFileName = input('Filename: ')
path = r"C:\Users\SahilPatil\Partners MGU\PBM Team - Documents\MRx Repricing"
ptps = load_workbook(filename=r"C:\Users\SahilPatil\Partners MGU\PBM Team - Documents\Data\DO NOT DELETE.xlsm", data_only = False)
#Import Magellan Reprice Information
Prospect = pd.read_excel(path + '\\' + MRxFileName + '.xlsx')
Prospect = Prospect.rename({'Unnamed: 0':'A','Unnamed: 1':'B','Unnamed: 2':'C','Unnamed: 3':'D','Unnamed: 4':'E','Unnamed: 5':'F','Unnamed: 6':'G','Unnamed: 7':'H','Unnamed: 8':'I','Unnamed: 9':'J','Unnamed: 10':'K','Unnamed: 11':'L','Unnamed: 12':'M','Unnamed: 13':'N','Unnamed: 14':'O'}, axis = 1)
#Get Sheet Names from PTPS
sheet_names = ptps.sheetnames
name = sheet_names[0]
terms = sheet_names[2]
Sheet1 = sheet_names[1]
paydhealth = sheet_names[3]
Output = sheet_names[4]
Sheet2 = sheet_names[5]
sheet_ranges = ptps[name]
sheet_ranges1 = ptps[terms]
sheet_ranges2 = ptps[Sheet1]
sheet_ranges3 = ptps[paydhealth]
sheet_ranges4 = ptps[Output]
sheet_ranges5 = ptps[Sheet2]

#Create Inputs df
Inputs = pd.DataFrame(sheet_ranges.values)
Terms = pd.DataFrame(sheet_ranges1.values)
Sheet1 = pd.DataFrame(sheet_ranges2.values)
paydhealth = pd.DataFrame(sheet_ranges3.values)
Output = pd.DataFrame(sheet_ranges4.values)
Sheet2 = pd.DataFrame(sheet_ranges5.values)

Inputs = Inputs.rename({0:'A',1:'B',2:'C',3:'D',4:'E',5:'F',6:'G',7:'H',8:'I'}, axis = 1)

#Import Group Name
Inputs.loc[0,'B'] = MRxFileName

#Extracting Rx quantities from case
mbq = Prospect['D'][7]
mgq = Prospect['D'][8]
msq = Prospect['D'][9]
mail_rx_total = mbq + mgq + msq

rbq = Prospect['D'][12]
rgq = Prospect['D'][13]
rsq = Prospect['D'][14]
retail_rx_total = rbq + rgq + rsq

r90bq = Prospect['D'][17]
r90gq = Prospect['D'][18]
r90sq = Prospect['D'][19]
retail_90_rx_total = r90bq + r90gq + r90sq

total_claims = mail_rx_total + retail_90_rx_total + retail_rx_total

#Add Rx Quantities in PTPS
Inputs.loc[5,'B'] = rbq
Inputs.loc[6,'B'] = r90bq
Inputs.loc[7,'B'] = rgq
Inputs.loc[8,'B'] = r90gq
Inputs.loc[9,'B'] = rsq + r90sq
Inputs.loc[18,'B'] = mbq
Inputs.loc[19,'B'] = mgq
Inputs.loc[20,'B'] = msq
Inputs.loc[1,'G'] = total_claims


#Extract Member Copays
copaymb = Prospect['E'][7]
copaymg = Prospect['E'][8]
copayms = Prospect['E'][9]
copay_mail_total = copaymb + copaymg + copayms

copayrb = Prospect['E'][12]
copayrg = Prospect['E'][13]
copayrs = Prospect['E'][14]
copay_retail_total = copayrb + copayrg + copayrs

copayr90b = Prospect['E'][17]
copayr90g = Prospect['E'][18]
copayr90s = Prospect['E'][19]
copay_retail_90_total = copayr90b + copayr90g + copayr90s

copay_total = copay_mail_total + copay_retail_90_total + copay_retail_total

#Add Copays to PTPS
Inputs.loc[13,'C'] = abs(copay_total)
Inputs.loc[13,'D'] = abs(copay_total)


#Extract Original Plan Paid
mbpaid = Prospect['F'][7]
mgpaid = Prospect['F'][8]
mspaid = Prospect['F'][9]
mail_plan_paid = mbpaid + mgpaid + mspaid

rbpaid = Prospect['F'][12]
rgpaid = Prospect['F'][13]
rspaid = Prospect['F'][14]
retail_plan_paid = rbpaid + rgpaid + rspaid

r90bpaid = Prospect['F'][17]
r90gpaid = Prospect['F'][18]
r90spaid = Prospect['F'][19]
retail_90_plan_paid = r90bpaid + r90gpaid + r90spaid

total_plan_paid = mail_plan_paid + retail_90_plan_paid + retail_plan_paid


#Add Original Plan Paid to PTPS
Inputs.loc[5,'C'] = rbpaid
Inputs.loc[6,'C'] = r90bpaid
Inputs.loc[7,'C'] = rgpaid
Inputs.loc[8,'C'] = r90gpaid
Inputs.loc[9,'C'] = rspaid + r90spaid
Inputs.loc[18,'C'] = mbpaid
Inputs.loc[19,'C'] = mgpaid
Inputs.loc[20,'C'] = mspaid

#Extract MRx Total Cost
MRx_mbcost = Prospect['H'][7]
MRx_mgcost = Prospect['H'][8]
MRx_mscost = Prospect['H'][9]
MRx_mail_cost = MRx_mbcost + MRx_mgcost + MRx_mscost

MRx_rbcost = Prospect['H'][12]
MRx_rgcost = Prospect['H'][13]
MRx_rscost = Prospect['H'][14]
MRx_retail_cost = MRx_rbcost + MRx_rgcost + MRx_rscost

MRx_r90bcost = Prospect['H'][17]
MRx_r90gcost = Prospect['H'][18]
MRx_r90scost = Prospect['H'][19]
MRx_retail_90_cost = MRx_r90bcost + MRx_r90gcost + MRx_r90scost

#Add MRx Total Cost to PTPS
Inputs.loc[5,'D'] = MRx_rbcost
Inputs.loc[6,'D'] = MRx_r90bcost
Inputs.loc[7,'D'] = MRx_rgcost
Inputs.loc[8,'D'] = MRx_r90gcost
Inputs.loc[9,'D'] = MRx_rscost + MRx_r90scost
Inputs.loc[18,'D'] = MRx_mbcost
Inputs.loc[19,'D'] = MRx_mgcost
Inputs.loc[20,'D'] = MRx_mscost

#Extract Dispensing Fees
MRx_rdispfee = Prospect['I'][15]

#Add Dispensing Fees to PTPS
Inputs.loc[10,'D'] = MRx_rdispfee

#Extract Admin Fees
MRx_Retail_Admin_Fee = Prospect['J'][15] + Prospect['J'][20]
MRx_Mail_Admin_Fee = Prospect['J'][10]

#Add Admin Fees to PTPS
Inputs.loc[12,'D'] = MRx_Retail_Admin_Fee
Inputs.loc[21,'D'] = MRx_Mail_Admin_Fee

#Extract Rebates
MRx_Rebates_Mail_Brand = Prospect['L'][7]
MRx_Rebates_Mail_Specialty = Prospect['L'][9]
MRx_Rebates_Retail_Brand = Prospect['L'][12]
MRx_Rebates_Retail_90_Brand = Prospect['L'][17]
MRx_Rebates_Retail_Specialty = Prospect['L'][14]
MRx_Rebtail_Retail_90_Specialty = Prospect['L'][19]

#Add Rebates to PTPS
Inputs.loc[14,'D'] = abs(MRx_Rebates_Retail_Brand)
Inputs.loc[15,'D'] = abs(MRx_Rebates_Retail_90_Brand)
Inputs.loc[16,'D'] = abs(MRx_Rebates_Retail_Specialty) + abs(MRx_Rebtail_Retail_90_Specialty)
Inputs.loc[22,'D'] = abs(MRx_Rebates_Mail_Brand)
Inputs.loc[23,'D'] = abs(MRx_Rebates_Mail_Specialty)

#Export
#Writer is supposed to export the Case Information that I update from Magellan's Reprice
writer = pd.ExcelWriter(r"C:\Users\SahilPatil\OneDrive - Partners MGU\Desktop\PBM Project\PyFiles\Case.xlsx", engine = 'xlsxwriter')

Inputs.to_excel(writer,header = None, index = False,sheet_name = 'Inputs')
Terms.to_excel(writer, sheet_name = 'Term Analysis', header = None, index = False)
Sheet1.to_excel(writer, sheet_name = 'Sheet1', header = None, index = False)
paydhealth.to_excel(writer, sheet_name = 'paydhealth', header = None, index = False)
Output.to_excel(writer, sheet_name = 'Output', header = None, index = False)
Sheet2.to_excel(writer, sheet_name = 'Sheet2', header = None, index = False)

writer.save()

